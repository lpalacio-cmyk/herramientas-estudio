"""
Procesador F931 (ARCA) — versión web
------------------------------------
Procesa PDFs del formulario 931 descargados de ARCA, extrae los datos
clave por OCR, y los muestra en una grilla editable que el usuario puede
corregir antes de descargar el Excel resumen.

Datos extraídos (cuando es posible):
- Período (de la URL del pie del PDF — confiable)
- CUIT, Razón Social, Empleados
- Suma de Rem. 1 y 9 (best effort por OCR)
- Sección VIII completa: 351, 301, 352, 302, 312, 028, 360, 270, 935
- Retenciones aplicadas a S.S. y O.S. (caso constructoras)

Si OCR falla en un dato, aparece como None y el usuario lo corrige
en la grilla en pantalla antes de descargar.
"""

import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st


# ============================================================
#  Lógica de extracción (OCR)
# ============================================================

def _imports():
    """Imports diferidos para que el resto de la app funcione si OCR no está disponible."""
    import pytesseract
    from pdf2image import convert_from_bytes
    from PIL import Image
    return pytesseract, convert_from_bytes, Image


def limpiar_numero(s):
    """Limpia errores OCR comunes en números formato AR."""
    if not s:
        return None
    s = re.sub(r"[^\d.,\-]", "", s.strip())
    s = s.replace(",.", ".")
    if not s:
        return None
    if re.match(r"^-?\d{1,3}(\.\d{3})*,\d{2}$", s) or re.match(r"^-?\d+,\d{2}$", s):
        return s
    if re.match(r"^-?\d{1,3}(,\d{3})*\.\d{2}$", s):
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    return None


def _buscar_cuit(textos, img_list):
    pytesseract, _, _ = _imports()
    # Buscar en texto OCR
    for txt in textos:
        m = re.search(r"(\d{2}-\d{8}-\d)", txt)
        if m:
            return m.group(1)
        m = re.search(r"(\d{2}\s+\d{8}\s+\d)", txt)
        if m:
            return m.group(1).replace(" ", "-")
    # Recortar zona del CUIT (top-right) y OCR específico
    for img in img_list:
        ancho, alto = img.size
        rec = img.crop((int(ancho * 0.55), 0, int(ancho * 0.99), int(alto * 0.06)))
        txt = pytesseract.image_to_string(rec, lang='eng')
        m = re.search(r"(\d{2}-\d{8}-\d)", txt)
        if m:
            return m.group(1)
    return None


def _buscar_empleados(texto):
    m = re.search(r"CUILES con ART\s+(\d+)", texto)
    if m:
        return int(m.group(1))
    idx = texto.find("CUILES con ART")
    if idx > 0:
        siguiente = texto[idx:idx + 500]
        m = re.search(r"\n(\d{1,4})\s+(\d{1,3}(?:\.\d{3})*,\d{2})", siguiente)
        if m:
            return int(m.group(1))
        m = re.search(r"\n(\d{1,4})\s*\n", siguiente)
        if m:
            return int(m.group(1))
    return None


def _extraer_codigos_viii(texto):
    """Extrae los 9 códigos del bloque VIII manejando layouts variables."""
    inicio = texto.find("MONTOS QUE SE INGRESAN")
    if inicio < 0:
        return {}
    bloque = texto[inicio:inicio + 2500]

    codigos = ["351", "301", "352", "302", "312", "028", "360", "270", "935"]
    resultado = {}

    # Estrategia 1: código y valor en la misma línea
    for cod in codigos:
        patron = rf"{cod}\s*[-)][^\n]*?(\d{{1,3}}(?:\.\d{{3}})*,\d{{2}})\s*$"
        m = re.search(patron, bloque, re.MULTILINE)
        if m:
            resultado[cod] = limpiar_numero(m.group(1))

    # Estrategia 2: códigos consecutivos sin valor inline (los valores vienen abajo)
    lineas = bloque.split("\n")
    for i, linea in enumerate(lineas):
        m_cod = re.match(r"^\s*(\d{3})\s*[-)]", linea)
        if not m_cod:
            continue
        cod = m_cod.group(1)
        if cod not in codigos or cod in resultado:
            continue
        if re.search(r"\d{1,3}(?:\.\d{3})*,\d{2}", linea):
            continue

        codigos_pendientes = [cod]
        for j in range(i + 1, min(i + 6, len(lineas))):
            l = lineas[j]
            otro_cod = re.match(r"^\s*(\d{3})\s*[-)]", l)
            if otro_cod and otro_cod.group(1) in codigos and otro_cod.group(1) not in resultado:
                if not re.search(r"\d{1,3}(?:\.\d{3})*,\d{2}", l):
                    codigos_pendientes.append(otro_cod.group(1))
                    continue
            montos = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", l)
            if montos and codigos_pendientes:
                for monto in montos:
                    if codigos_pendientes:
                        c = codigos_pendientes.pop(0)
                        if c not in resultado:
                            resultado[c] = limpiar_numero(monto)
                if not codigos_pendientes:
                    break

    return resultado


def _extraer_sumas_rem(img):
    """OCR fila por fila del cuadro de Sumas de Rem."""
    pytesseract, _, Image = _imports()
    ancho, alto = img.size
    x1 = int(ancho * 0.83)
    x2 = int(ancho * 0.99)
    y_inicio = int(alto * 0.075)
    y_fin = int(alto * 0.218)
    alto_fila = (y_fin - y_inicio) / 10

    sumas = {}
    for i in range(10):
        n = i + 1
        y_top = max(0, int(y_inicio + i * alto_fila) - 3)
        y_bot = min(alto, int(y_inicio + (i + 1) * alto_fila) + 3)
        fila = img.crop((x1, y_top, x2, y_bot))
        gris = fila.convert("L")
        escalado = gris.resize((gris.width * 3, gris.height * 3), Image.LANCZOS)
        txt = pytesseract.image_to_string(escalado, lang='eng', config='--psm 7').strip()

        m = re.search(r"\d{1,3}(?:[.,]\d{3})+[.,]\d{2}", txt)
        if m:
            sumas[n] = limpiar_numero(m.group(0))
        elif re.search(r"\b0[,.]00\b", txt) or txt.strip() in ("0", "0)", "0]", "0|"):
            sumas[n] = "0,00"
        else:
            sumas[n] = None
    return sumas


def extraer_datos_f931(pdf_bytes: bytes) -> dict:
    """Extrae todos los datos de un PDF F931. Devuelve dict con valores o None si no detectó."""
    pytesseract, convert_from_bytes, _ = _imports()

    resultado = {
        "periodo": None, "cuit": None, "razon_social": None,
        "empleados": None, "rem_1": None, "rem_9": None,
        "351": None, "301": None, "352": None, "302": None,
        "312": None, "028": None, "360": None, "270": None, "935": None,
        "ret_ss": "0,00", "ret_os": "0,00",
        "errores": []
    }

    try:
        img300 = convert_from_bytes(pdf_bytes, dpi=300)[0]
        img400 = convert_from_bytes(pdf_bytes, dpi=400)[0]
    except Exception as e:
        resultado["errores"].append(f"PDF inválido: {e}")
        return resultado

    texto300 = pytesseract.image_to_string(img300, lang='eng')
    texto400 = pytesseract.image_to_string(img400, lang='eng')

    # Período (URL del pie)
    for txt in (texto300, texto400):
        m = re.search(r"Periodo=(\d{4})(\d{2})", txt)
        if m:
            resultado["periodo"] = f"{m.group(2)}/{m.group(1)}"
            break
    if not resultado["periodo"]:
        resultado["errores"].append("Período no detectado")

    # CUIT
    cuit = _buscar_cuit([texto300, texto400], [img300, img400])
    if cuit:
        resultado["cuit"] = cuit

    # Razón Social
    for txt in (texto300, texto400):
        m = re.search(
            r"Raz[oó]n\s+Social:?[\s\S]{1,80}\n\n([A-ZÑ][A-ZÑ\.\s&\d]+?(?:S\.?R\.?L\.?|S\.?A\.?S?|SA|SRL|SAS))",
            txt
        )
        if m:
            resultado["razon_social"] = re.sub(r"\s+", " ", m.group(1)).strip()
            break

    # Empleados
    for txt in (texto300, texto400):
        e = _buscar_empleados(txt)
        if e:
            resultado["empleados"] = e
            break

    # Sumas de Rem
    try:
        sumas = _extraer_sumas_rem(img400)
        resultado["rem_1"] = sumas.get(1)
        resultado["rem_9"] = sumas.get(9)
    except Exception as e:
        resultado["errores"].append(f"Sumas Rem: {e}")

    # Códigos VIII (combinar textos para mejor cobertura)
    cods = _extraer_codigos_viii(texto300)
    cods400 = _extraer_codigos_viii(texto400)
    for k, v in cods400.items():
        if cods.get(k) is None:
            cods[k] = v
    for cod in ["351", "301", "352", "302", "312", "028", "360", "270", "935"]:
        resultado[cod] = cods.get(cod)

    # Retenciones
    for txt in (texto300, texto400):
        if resultado["ret_ss"] == "0,00":
            m = re.search(r"Retenciones aplicadas a Seguridad Social\s+(\d{1,3}(?:\.\d{3})*,\d{2}|0[,.]00)", txt)
            if m:
                resultado["ret_ss"] = limpiar_numero(m.group(1)) or "0,00"
        if resultado["ret_os"] == "0,00":
            m = re.search(r"Retenciones aplicadas a Obra Social\s+(\d{1,3}(?:\.\d{3})*,\d{2}|0[,.]00)", txt)
            if m:
                resultado["ret_os"] = limpiar_numero(m.group(1)) or "0,00"

    return resultado


# ============================================================
#  Helpers para la grilla
# ============================================================

CONCEPTOS_GRILLA = [
    ("Empleados", "empleados", "entero"),
    ("Suma de Rem. 1", "rem_1", "monto"),
    ("Suma de Rem. 9", "rem_9", "monto"),
    ("351 - Contribuciones SS", "351", "monto"),
    ("301 - Aportes SS", "301", "monto"),
    ("352 - Contribuciones OS", "352", "monto"),
    ("302 - Aportes OS", "302", "monto"),
    ("312 - L.R.T.", "312", "monto"),
    ("028 - Seg. Colectivo Vida", "028", "monto"),
    ("360 - Contribuciones RENATRE", "360", "monto"),
    ("270 - Vales Alimentarios", "270", "monto"),
    ("935 - Seg. Sepelio UATRE", "935", "monto"),
    ("Retenciones aplicadas a SS", "ret_ss", "monto"),
    ("Retenciones aplicadas a OS", "ret_os", "monto"),
]

CODIGOS_TOTAL_VIII = ["351", "301", "352", "302", "312", "028", "360", "270", "935"]


def parsear_monto_ar(s):
    """Convierte '5.552.957,18' a float. Devuelve None si no se puede."""
    if s is None or s == "" or s == "REVISAR":
        return None
    s = str(s).strip()
    if not s:
        return None
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def construir_grilla(resultados_por_pdf):
    """
    Construye DataFrame con conceptos en filas y meses en columnas.
    Ordena los meses cronológicamente. Calcula el Total VIII.
    """
    # Filtrar PDFs que tienen período (sin período no se puede ubicar la columna)
    validos = [r for r in resultados_por_pdf if r.get("periodo")]
    # Ordenar por período (MM/YYYY → ordenar por año, luego mes)
    def clave(r):
        mm, yy = r["periodo"].split("/")
        return (int(yy), int(mm))
    validos.sort(key=clave)

    columnas = [r["periodo"] for r in validos]
    filas = []

    # Conceptos directos
    for etiqueta, campo, _ in CONCEPTOS_GRILLA:
        fila = {"Concepto": etiqueta}
        for r in validos:
            v = r.get(campo)
            fila[r["periodo"]] = "REVISAR" if v is None else str(v)
        filas.append(fila)

    # Total VIII calculado (al final)
    fila_total = {"Concepto": "TOTAL VIII (suma)"}
    for r in validos:
        suma = 0.0
        algun_revisar = False
        for cod in CODIGOS_TOTAL_VIII:
            v = parsear_monto_ar(r.get(cod))
            if v is None:
                algun_revisar = True
            else:
                suma += v
        if algun_revisar:
            fila_total[r["periodo"]] = "REVISAR"
        else:
            fila_total[r["periodo"]] = f"{suma:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    filas.append(fila_total)

    df = pd.DataFrame(filas)
    return df, columnas


def grilla_a_excel(df: pd.DataFrame, info_cliente: dict) -> bytes:
    """Genera Excel con la grilla y un encabezado con datos del cliente."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "F931 Resumen"

    font_titulo = Font(bold=True, size=12)
    font_hdr = Font(bold=True, size=10, color="FFFFFF")
    fill_hdr = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    fill_total = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
    align_c = Alignment(horizontal="center", vertical="center")

    # Encabezado
    ws.cell(row=1, column=1, value="Resumen F.931").font = font_titulo
    ws.cell(row=2, column=1, value=f"Cliente: {info_cliente.get('razon_social', '—')}")
    ws.cell(row=3, column=1, value=f"CUIT: {info_cliente.get('cuit', '—')}")
    ws.cell(row=4, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    # Fila de encabezados de la tabla (fila 6)
    fila_hdr = 6
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=fila_hdr, column=j, value=col)
        c.font = font_hdr
        c.fill = fill_hdr
        c.alignment = align_c

    # Datos
    for i, row in df.iterrows():
        fila_excel = fila_hdr + 1 + i
        es_total = (row["Concepto"] == "TOTAL VIII (suma)")
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            cell = ws.cell(row=fila_excel, column=j)
            if j == 1:
                # Columna de concepto
                cell.value = v
                if es_total:
                    cell.font = Font(bold=True)
            else:
                # Columnas de meses
                if v == "REVISAR" or v is None or v == "":
                    cell.value = v if v else None
                else:
                    # Intentar convertir a número
                    try:
                        n = float(str(v).replace(".", "").replace(",", "."))
                        cell.value = n
                        cell.number_format = "#,##0.00"
                    except (ValueError, AttributeError):
                        cell.value = v
            if es_total:
                cell.fill = fill_total

    # Anchos
    ws.column_dimensions["A"].width = 32
    for j in range(2, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16

    ws.freeze_panes = "B7"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
#  Interfaz
# ============================================================

st.title("📄 Procesador F.931")
st.markdown(
    "Subí los PDFs del formulario 931 descargados de ARCA (uno por mes) y "
    "obtené un resumen consolidado con todos los datos clave organizados "
    "por período. Podés corregir valores en pantalla antes de descargar."
)

with st.expander("¿Cómo lo uso?"):
    st.markdown(
        """
        1. Descargá de ARCA los F.931 de los meses que necesités. Pueden tener
           cualquier nombre (`01-25.pdf`, `enero.pdf`, etc.) — el período se
           detecta del propio PDF.
        2. Subí todos los PDFs juntos abajo.
        3. Apretá **Procesar**. Tarda ~5-10 segundos por PDF (es OCR).
        4. Revisá la grilla. Las celdas con `REVISAR` son las que el OCR no
           pudo leer con seguridad — completalas vos manualmente.
        5. Cuando esté todo bien, podés:
           - **Copiar y pegar la grilla** en tu papel de trabajo, o
           - **Descargar el Excel** consolidado.

        💡 **Limitación honesta**: las "Suma de Rem. 1" y "Suma de Rem. 9" 
        están en un cuadro chico arriba a la derecha del PDF que el OCR 
        a veces lee mal. Si aparece `REVISAR`, completalos manualmente. 
        Para el resto, el OCR es muy confiable.
        """
    )

archivos = st.file_uploader(
    "PDFs del F.931",
    type=["pdf"],
    accept_multiple_files=True,
    help="Podés subir hasta 12 PDFs juntos (uno por mes).",
)

if archivos:
    st.caption(f"📑 {len(archivos)} archivo(s) cargado(s).")

# Estado en sesión: guardamos los resultados para que la grilla sea editable
if "f931_resultados" not in st.session_state:
    st.session_state.f931_resultados = None

if st.button("Procesar PDFs", type="primary", disabled=not archivos):
    resultados = []
    progreso = st.progress(0, text="Procesando...")
    errores_globales = []

    for i, archivo in enumerate(archivos):
        progreso.progress((i + 1) / len(archivos), text=f"OCR de {archivo.name}...")
        try:
            datos = extraer_datos_f931(archivo.getvalue())
            datos["_archivo"] = archivo.name
            resultados.append(datos)
            if datos["errores"]:
                errores_globales.append((archivo.name, datos["errores"]))
        except Exception as e:
            errores_globales.append((archivo.name, [f"Error fatal: {e}"]))

    progreso.empty()
    st.session_state.f931_resultados = resultados

    if errores_globales:
        with st.expander("⚠️ Avisos del procesamiento", expanded=False):
            for nombre, errs in errores_globales:
                st.markdown(f"**{nombre}**: {', '.join(errs)}")

# ============================================================
#  Mostrar grilla editable
# ============================================================

if st.session_state.f931_resultados:
    resultados = st.session_state.f931_resultados
    validos = [r for r in resultados if r.get("periodo")]
    invalidos = [r for r in resultados if not r.get("periodo")]

    if invalidos:
        with st.warning(""):
            pass
        st.warning(
            f"⚠️ {len(invalidos)} archivo(s) sin período detectable, no se incluyeron: "
            + ", ".join(r["_archivo"] for r in invalidos)
        )

    if not validos:
        st.error("Ningún PDF se pudo procesar. Verificá que sean F.931 válidos.")
        st.stop()

    # Datos de cliente (tomar del primer PDF que los tenga)
    razon = next((r["razon_social"] for r in validos if r.get("razon_social")), "—")
    cuit_detectado = next((r["cuit"] for r in validos if r.get("cuit")), None)

    st.divider()
    st.subheader("Datos del cliente")
    col_a, col_b = st.columns(2)
    with col_a:
        st.text_input("Razón Social", value=razon, disabled=True)
    with col_b:
        cuit_input = st.text_input(
            "CUIT",
            value=cuit_detectado or "",
            placeholder="30-XXXXXXXX-X",
            help="Si quedó vacío, completalo manualmente.",
        )

    # Construir grilla
    df_grilla, columnas_meses = construir_grilla(validos)

    st.divider()
    st.subheader("Datos extraídos por mes")

    # Mostrar resumen de cobertura
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("PDFs procesados", len(validos))
    with col2:
        # Contar celdas con REVISAR
        a_revisar = (df_grilla.iloc[:, 1:] == "REVISAR").sum().sum()
        st.metric("Celdas a revisar", int(a_revisar))
    with col3:
        total_celdas = (df_grilla.shape[0] - 1) * (df_grilla.shape[1] - 1)
        cobertura = (1 - a_revisar / total_celdas) * 100 if total_celdas else 0
        st.metric("Cobertura OCR", f"{cobertura:.0f}%")

    st.markdown(
        "**Editá la grilla**: hacé clic en una celda con `REVISAR` y escribí "
        "el valor correcto. Después podés copiar y pegar (`Ctrl+C` / `Ctrl+V`) "
        "directo a tu papel de trabajo, o descargar el Excel."
    )

    # Configuración de columnas: todas TextColumn para aceptar cualquier formato
    column_config = {"Concepto": st.column_config.TextColumn("Concepto", disabled=True, width="large")}
    for col in columnas_meses:
        column_config[col] = st.column_config.TextColumn(
            col,
            help=f"Datos del período {col}",
            width="small",
        )

    df_editado = st.data_editor(
        df_grilla,
        column_config=column_config,
        hide_index=True,
        use_container_width=True,
        key="grilla_f931",
    )

    # Recalcular Total VIII en vivo basado en lo editado
    st.divider()
    st.subheader("Validación")

    fila_total_idx = len(df_editado) - 1  # última fila
    discrepancias = []
    for col in columnas_meses:
        # Sumar los códigos del bloque VIII
        suma = 0.0
        algun_invalido = False
        for codigo in CODIGOS_TOTAL_VIII:
            # Encontrar la fila del código
            etiqueta = next((e for e, c, _ in CONCEPTOS_GRILLA if c == codigo), None)
            if etiqueta:
                v = df_editado.loc[df_editado["Concepto"] == etiqueta, col].iloc[0]
                n = parsear_monto_ar(v)
                if n is None:
                    algun_invalido = True
                else:
                    suma += n

        # Comparar contra el valor de Total VIII en la grilla
        v_total = df_editado.loc[df_editado["Concepto"] == "TOTAL VIII (suma)", col].iloc[0]
        n_total = parsear_monto_ar(v_total)

        if not algun_invalido and n_total is not None:
            if abs(suma - n_total) > 0.01:
                discrepancias.append((col, suma, n_total))

    if discrepancias:
        st.warning(
            "⚠️ Hay diferencias entre el Total VIII calculado y el de la grilla "
            "(probablemente edités algún código pero no actualizaste el total):"
        )
        for col, calc, mostrado in discrepancias:
            st.caption(f"   {col}: calculado {calc:,.2f}  vs grilla {mostrado:,.2f}")
    else:
        if a_revisar == 0:
            st.success("✅ Todos los datos completos y consistentes.")
        else:
            st.info("ℹ️ Hay celdas en `REVISAR`. Completalas para validar los totales.")

    # Descarga
    st.divider()
    st.subheader("Descargar Excel")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"F931_Resumen_{timestamp}.xlsx"

    info_cliente = {"razon_social": razon, "cuit": cuit_input}
    excel_bytes = grilla_a_excel(df_editado, info_cliente)

    st.download_button(
        label=f"⬇️ Descargar {nombre_archivo}",
        data=excel_bytes,
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    st.caption(
        "💡 También podés seleccionar las celdas de la grilla de arriba con el "
        "mouse y copiar (`Ctrl+C`) para pegar directo en tu papel de trabajo de Excel."
    )
