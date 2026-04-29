"""
Procesador Libro IVA — versión web
----------------------------------
Conversión del script consolidar_iva.py a una página de Streamlit.
La lógica de extracción y formateo es la misma; solo cambian las E/S:
  - Entrada: archivos ZIP subidos por el usuario (en vez de una carpeta).
  - Salida: dos Excel descargables (en vez de guardarse a disco).
"""

import io
import zipfile
from typing import Optional

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
#  Lógica de negocio (idéntica al script original)
# ============================================================

def leer_csv_de_zip(zip_bytes: bytes, nombre_zip: str, nombre_csv: str) -> pd.DataFrame:
    """Lee un CSV específico desde un ZIP en memoria. Devuelve DataFrame vacío si no está."""
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            nombres = [n for n in z.namelist() if n.lower().endswith(nombre_csv.lower())]
            if not nombres:
                return pd.DataFrame()
            with z.open(nombres[0]) as f:
                raw = f.read()
    except Exception as e:
        st.warning(f"No se pudo leer {nombre_zip}: {e}")
        return pd.DataFrame()

    for enc in ("utf-8-sig", "iso-8859-1", "latin1"):
        try:
            df = pd.read_csv(
                io.BytesIO(raw),
                sep=";",
                encoding=enc,
                dtype=str,
                skip_blank_lines=True,
            )
            df.columns = [c.strip().strip('"') for c in df.columns]
            return df
        except Exception:
            continue
    return pd.DataFrame()


def consolidar(zips_subidos: list, nombre_csv: str) -> tuple[pd.DataFrame, list]:
    """
    Consolida un CSV específico desde todos los ZIPs subidos.
    Devuelve el DataFrame consolidado y un log con el detalle por ZIP.
    """
    frames = []
    log = []

    for archivo in zips_subidos:
        df = leer_csv_de_zip(archivo.getvalue(), archivo.name, nombre_csv)
        if not df.empty:
            log.append(("ok", archivo.name, len(df)))
            frames.append(df)
        else:
            log.append(("vacio", archivo.name, 0))

    if not frames:
        return pd.DataFrame(), log

    consolidado = pd.concat(frames, ignore_index=True)
    return consolidado, log


def exportar_excel(df: pd.DataFrame, titulo: str) -> bytes:
    """Genera el Excel formateado y lo devuelve como bytes (en memoria)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    COLOR_HEADER = "1F4E79"
    font_hdr = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    font_dat = Font(name="Arial", size=10)
    fill_hdr = PatternFill("solid", start_color=COLOR_HEADER, end_color=COLOR_HEADER)
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Encabezados
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = align_c
    ws.row_dimensions[1].height = 32

    cols_num = {
        col for col in df.columns
        if any(kw in col for kw in ["Importe", "Neto", "Total", "Crédito", "Tipo de Cambio"])
    }
    cols_fecha = {col for col in df.columns if "Fecha" in col}

    # Datos
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            col_name = df.columns[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_dat

            if pd.isna(value) if not isinstance(value, str) else str(value).strip() in ("", "nan"):
                cell.value = None
            else:
                if col_name in cols_num:
                    try:
                        cell.value = float(str(value).replace(",", "."))
                        cell.number_format = "#,##0.00"
                    except ValueError:
                        cell.value = value
                elif col_name in cols_fecha:
                    cell.value = value
                    cell.number_format = "DD/MM/YYYY"
                else:
                    cell.value = value

    # Anchos de columna
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if "Denominación" in col_name:
            ws.column_dimensions[col_letter].width = 32
        elif "Fecha" in col_name:
            ws.column_dimensions[col_letter].width = 14
        elif col_name in cols_num:
            ws.column_dimensions[col_letter].width = 16
        else:
            ws.column_dimensions[col_letter].width = max(10, min(len(col_name), 20))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Guardar a bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
#  Interfaz de usuario
# ============================================================

st.title("📑 Procesador Libro IVA")
st.markdown(
    "Subí los archivos ZIP descargados de AFIP. La herramienta extrae "
    "`comprobantes_ventas.csv` y `comprobantes_compras.csv` de cada uno y "
    "genera dos Excel consolidados con el formato listo para pegar."
)

with st.expander("¿Cómo lo uso?"):
    st.markdown(
        """
        1. Descargá de AFIP los libros IVA de los meses que necesités. Cada mes
           queda en un ZIP.
        2. Arrastrá todos los ZIPs (o seleccionalos) en el cuadro de abajo. Podés
           subir uno o muchos a la vez.
        3. Apretá **Procesar**. Cuando termine, vas a poder descargar
           `IVA_ventas.xlsx` y `IVA_compras.xlsx`.

        No importa el orden ni cómo se llamen los archivos: la herramienta busca
        adentro de cada ZIP los CSV correspondientes.
        """
    )

archivos = st.file_uploader(
    "ZIPs del libro IVA",
    type=["zip"],
    accept_multiple_files=True,
    help="Podés seleccionar varios a la vez (Ctrl+clic) o arrastrarlos todos juntos.",
)

if archivos:
    st.caption(f"📦 {len(archivos)} archivo(s) cargado(s).")

procesar = st.button("Procesar", type="primary", disabled=not archivos)

if procesar:
    with st.spinner("Procesando ZIPs..."):
        df_ventas, log_ventas = consolidar(archivos, "comprobantes_ventas.csv")
        df_compras, log_compras = consolidar(archivos, "comprobantes_compras.csv")

    st.divider()

    # Detalle de procesamiento
    with st.expander("Detalle por archivo", expanded=False):
        col_v, col_c = st.columns(2)
        with col_v:
            st.markdown("**Ventas**")
            for estado, nombre, filas in log_ventas:
                if estado == "ok":
                    st.markdown(f"✅ `{nombre}` — {filas} filas")
                else:
                    st.markdown(f"➖ `{nombre}` — sin CSV de ventas")
        with col_c:
            st.markdown("**Compras**")
            for estado, nombre, filas in log_compras:
                if estado == "ok":
                    st.markdown(f"✅ `{nombre}` — {filas} filas")
                else:
                    st.markdown(f"➖ `{nombre}` — sin CSV de compras")

    # Resultado
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Ventas")
        if not df_ventas.empty:
            st.metric("Filas consolidadas", len(df_ventas))
            xlsx_ventas = exportar_excel(df_ventas, "Ventas")
            st.download_button(
                label="⬇️ Descargar IVA_ventas.xlsx",
                data=xlsx_ventas,
                file_name="IVA_ventas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            with st.expander("Vista previa"):
                st.dataframe(df_ventas.head(50), use_container_width=True)
        else:
            st.info("Ningún ZIP contenía `comprobantes_ventas.csv`.")

    with col2:
        st.subheader("Compras")
        if not df_compras.empty:
            st.metric("Filas consolidadas", len(df_compras))
            xlsx_compras = exportar_excel(df_compras, "Compras")
            st.download_button(
                label="⬇️ Descargar IVA_compras.xlsx",
                data=xlsx_compras,
                file_name="IVA_compras.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            with st.expander("Vista previa"):
                st.dataframe(df_compras.head(50), use_container_width=True)
        else:
            st.info("Ningún ZIP contenía `comprobantes_compras.csv`.")

    if not df_ventas.empty or not df_compras.empty:
        st.success("Proceso completado.")
