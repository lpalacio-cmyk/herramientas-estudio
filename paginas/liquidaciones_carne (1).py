"""
paginas/liquidaciones_carne.py

Procesador de liquidaciones de compra de carne (LCD, LCDP, LC) descargadas
de ARCA. Extrae datos de los PDFs y arma el Excel consolidado en el formato
del estudio.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timedelta, date
from decimal import Decimal

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# === REGEX =================================================================

RE_CPTE = re.compile(r"N°\s*(\d{4,5}-\d{8})")
RE_FECHA = re.compile(r"Fecha\s+(\d{2}/\d{2}/\d{4})")
RE_IMPORTE_BRUTO = re.compile(r"Importe\s+Bruto:\s*\$\s*([\d,]+\.\d{2})")
RE_TOTAL_GASTOS = re.compile(r"Total\s+Gastos:\s*\$\s*([\d,]+\.\d{2})?")
RE_FILA_CAT = re.compile(
    r"Kg\.\s*Vivo\s+([\d,]+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+(\d+\.\d{2})"
)
RE_AJUSTE = re.compile(r"Ajuste\s+Físico\s+de\s+Crédito", re.IGNORECASE)
RE_LCDP = re.compile(
    r"Liquidaci[óo]n\s+Compra\s+Directa.*Porcinos", re.IGNORECASE | re.DOTALL
)
RE_LCD = re.compile(r"Liquidaci[óo]n\s+Compra\s+Directa", re.IGNORECASE)
RE_LC = re.compile(r"Liquidaci[óo]n\s+de\s+compra(?!\s+Directa)", re.IGNORECASE)


# === HELPERS ===============================================================

def _to_decimal(s: str) -> Decimal:
    return Decimal(str(s).replace(",", ""))


def _ultimo_dia_mes(d: date) -> date:
    if d.month == 12:
        return d.replace(day=31)
    primer_dia_siguiente = d.replace(day=1, month=d.month + 1)
    return primer_dia_siguiente - timedelta(days=1)


def detectar_tipo(texto: str) -> tuple[str | None, bool]:
    es_ajuste = bool(RE_AJUSTE.search(texto))
    cabecera = "\n".join(texto.splitlines()[:30])
    if RE_LC.search(cabecera) and not RE_LCD.search(cabecera):
        return "LC", es_ajuste
    if RE_LCDP.search(cabecera):
        return "LCDP", es_ajuste
    if RE_LCD.search(cabecera):
        return "LCD", es_ajuste
    return None, es_ajuste


# === PARSER ================================================================

def parsear_pdf_bytes(file_bytes: bytes, nombre_archivo: str) -> dict:
    """Parsea un PDF de liquidacion y devuelve dict con datos + auditoria."""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        texto = pdf.pages[0].extract_text() or ""

    tipo, es_ajuste = detectar_tipo(texto)

    m_cpte = RE_CPTE.search(texto)
    cpte = m_cpte.group(1) if m_cpte else None

    m_fecha = RE_FECHA.search(texto)
    try:
        fecha = (
            datetime.strptime(m_fecha.group(1), "%d/%m/%Y").date()
            if m_fecha
            else None
        )
    except (AttributeError, ValueError):
        fecha = None

    mes = _ultimo_dia_mes(fecha) if fecha else None

    # Filas de categoria/raza
    filas = RE_FILA_CAT.findall(texto)
    cantidades = [int(c.replace(",", "")) for c, _, _, _ in filas]
    brutos_categoria = [float(_to_decimal(b)) for _, _, b, _ in filas]

    cantidad_total = sum(cantidades)
    bruto_categorias_total = sum(brutos_categoria)

    # Importe Bruto desde totales (sanity check)
    m_ib = RE_IMPORTE_BRUTO.search(texto)
    importe_bruto_pdf = float(_to_decimal(m_ib.group(1))) if m_ib else None

    # Total Gastos (en LC es la comision)
    total_gastos = 0.0
    m_tg = RE_TOTAL_GASTOS.search(texto)
    if m_tg and m_tg.group(1):
        total_gastos = float(_to_decimal(m_tg.group(1)))

    # Mercaderia: usamos Importe Bruto del total (fuente de verdad);
    # si no aparece, sumamos las filas
    bruto_mercaderia = (
        importe_bruto_pdf if importe_bruto_pdf is not None else bruto_categorias_total
    )
    bruto_final = bruto_mercaderia + total_gastos

    # Si es ajuste, invertir signo
    signo = -1 if es_ajuste else 1
    cantidad_total *= signo
    bruto_final *= signo
    bruto_mercaderia_signed = bruto_mercaderia * signo
    total_gastos_signed = total_gastos * signo

    # Componentes para construir formulas en el Excel
    componentes_kg = [c * signo for c in cantidades]
    componentes_bruto = [b * signo for b in brutos_categoria]
    if total_gastos > 0:
        componentes_bruto.append(total_gastos * signo)

    # Validaciones
    motivos = []
    if tipo is None:
        motivos.append("TIPO no detectado")
    if cpte is None:
        motivos.append("CPTE no detectado")
    if fecha is None:
        motivos.append("FECHA no detectada")
    if not filas:
        motivos.append("Sin filas de categoría")
    if (
        importe_bruto_pdf is not None
        and abs(importe_bruto_pdf - bruto_categorias_total) > 0.01
    ):
        motivos.append(
            f"Importe Bruto PDF ({importe_bruto_pdf:,.2f}) ≠ "
            f"suma de filas ({bruto_categorias_total:,.2f})"
        )
    if tipo == "LC" and total_gastos == 0:
        motivos.append("LC sin Comisión detectada")

    estado = "OK" if not motivos else "⚠️ REVISAR"

    return {
        "Estado": estado,
        "TIPO": tipo or "",
        "CPTE": cpte or "",
        "FECHA": fecha,
        "MES": mes,
        "CANTIDAD KG": cantidad_total,
        "$ BRUTO": round(bruto_final, 2),
        "$ Mercadería": round(bruto_mercaderia_signed, 2),
        "$ Comisión": round(total_gastos_signed, 2),
        "Filas": len(filas),
        "Archivo": nombre_archivo,
        "_componentes_kg": componentes_kg,
        "_componentes_bruto": componentes_bruto,
        "_es_ajuste": es_ajuste,
        "_motivos": "; ".join(motivos),
    }


# === EXPORT EXCEL ==========================================================

def _construir_formula(componentes: list, decimales: bool) -> object:
    """
    Si hay un solo componente devuelve el valor numerico.
    Si hay varios, devuelve fórmula tipo '=a+b' o '=-(a+b)' para multi negativo.
    """
    if not componentes:
        return None

    fmt = "{:.2f}" if decimales else "{:.0f}"

    if len(componentes) == 1:
        return componentes[0]

    # Detectamos si todos son negativos (caso ajuste con multi-fila)
    es_negativo = all(c < 0 for c in componentes)
    abs_componentes = [abs(c) for c in componentes]
    suma_str = "+".join(fmt.format(c) for c in abs_componentes)

    if es_negativo:
        return f"=-({suma_str})"
    return f"={suma_str}"


def generar_excel(filas: list[dict], incluir_archivo: bool = False) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    headers = ["TIPO", "CPTE", "FECHA", "MES", "CANTIDAD KG", "$ BRUTO"]
    if incluir_archivo:
        headers.append("Archivo")

    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Arial")

    for i, fila in enumerate(filas, start=2):
        ws.cell(row=i, column=1, value=fila["TIPO"]).font = body_font
        ws.cell(row=i, column=2, value=fila["CPTE"]).font = body_font
        ws.cell(row=i, column=3, value=fila["FECHA"]).font = body_font
        ws.cell(row=i, column=4, value=fila["MES"]).font = body_font

        kg = _construir_formula(fila["_componentes_kg"], decimales=False)
        ws.cell(row=i, column=5, value=kg).font = body_font

        bruto = _construir_formula(fila["_componentes_bruto"], decimales=True)
        ws.cell(row=i, column=6, value=bruto).font = body_font

        if incluir_archivo:
            ws.cell(row=i, column=7, value=fila["Archivo"]).font = body_font

    # Anchos
    widths = [8, 20, 12, 12, 14, 18]
    if incluir_archivo:
        widths.append(40)
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = w

    # Formatos
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "DD/MM/YYYY"
        ws.cell(row=row_idx, column=4).number_format = "DD/MM/YYYY"
        ws.cell(row=row_idx, column=5).number_format = "#,##0"
        ws.cell(row=row_idx, column=6).number_format = "#,##0.00"

    # Freeze header
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# === STREAMLIT UI ==========================================================

st.title("🥩 Procesador Liquidaciones de Compra de Carne")

with st.expander("ℹ️ Cómo funciona", expanded=False):
    st.markdown(
        """
        Subí PDFs de liquidaciones (LCD, LCDP, LC) descargados de ARCA y la
        herramienta extrae automáticamente:
        
        - **TIPO**: detectado del título del comprobante.
        - **CPTE**: número completo formato `XXXXX-XXXXXXXX`.
        - **FECHA** y **MES** (último día del mes).
        - **CANTIDAD KG**: suma de la columna "Cantidad" del cuadro Categoría / Raza.
        - **$ BRUTO**: Importe Bruto + Total Gastos (la comisión, en LC).
        
        **Reglas especiales:**
        - **Ajustes Físicos de Crédito** se cargan en negativo automáticamente.
        - **Multi-fila por raza** (ej. Toro + Vaca): se suman los valores.
        - **LC con comisión**: la comisión se suma al $ BRUTO.
        
        **Fórmulas vivas en el Excel**: cuando hay suma (multi-fila o comisión),
        la celda contiene la fórmula con sus componentes (ej. `=6627400+9882600`)
        para que puedas auditar haciendo clic en la celda.
        
        Las celdas marcadas **⚠️ REVISAR** indican algo que el extractor no pudo
        leer con seguridad y que conviene revisar manualmente.
        """
    )

# Estado
if "lc_filas" not in st.session_state:
    st.session_state.lc_filas = []

archivos = st.file_uploader(
    "Subí los PDFs de liquidaciones",
    type=["pdf"],
    accept_multiple_files=True,
    key="lc_uploader",
    help="Podés subir varios PDFs a la vez. Los duplicados (mismo CPTE) se omiten.",
)

col1, col2, _ = st.columns([2, 2, 6])
with col1:
    procesar = st.button(
        "📥 Procesar PDFs",
        type="primary",
        disabled=not archivos,
        use_container_width=True,
    )
with col2:
    if st.button("🗑️ Limpiar grilla", use_container_width=True):
        st.session_state.lc_filas = []
        st.rerun()

# Procesamiento
if procesar and archivos:
    nuevas: list[dict] = []
    cptes_existentes = {f["CPTE"] for f in st.session_state.lc_filas if f.get("CPTE")}
    duplicados: list[str] = []

    progress = st.progress(0, text="Procesando...")
    for i, archivo in enumerate(archivos):
        try:
            fila = parsear_pdf_bytes(archivo.read(), archivo.name)
            if fila["CPTE"] and fila["CPTE"] in cptes_existentes:
                duplicados.append(fila["CPTE"])
            else:
                nuevas.append(fila)
                if fila["CPTE"]:
                    cptes_existentes.add(fila["CPTE"])
        except Exception as e:
            st.error(f"❌ Error procesando {archivo.name}: {type(e).__name__}: {e}")
        progress.progress((i + 1) / len(archivos), text=f"Procesando... {i + 1}/{len(archivos)}")
    progress.empty()

    st.session_state.lc_filas.extend(nuevas)

    if nuevas:
        st.success(f"✅ {len(nuevas)} comprobante(s) procesados.")
    if duplicados:
        st.warning(
            f"⚠️ Se omitieron {len(duplicados)} duplicado(s) "
            f"(CPTE ya estaba en la grilla): {', '.join(duplicados)}"
        )

# Grilla
if st.session_state.lc_filas:
    # Ordenar por FECHA y CPTE
    filas_ordenadas = sorted(
        st.session_state.lc_filas,
        key=lambda f: (f["FECHA"] or date(1900, 1, 1), f.get("CPTE") or ""),
    )

    revisar_count = sum(1 for f in filas_ordenadas if f["Estado"] != "OK")
    if revisar_count:
        st.warning(
            f"⚠️ **{revisar_count}** comprobante(s) requieren revisión. "
            f"Mirá la columna **Estado** y editá si hace falta."
        )
        with st.expander("Ver detalle de motivos"):
            for f in filas_ordenadas:
                if f["Estado"] != "OK":
                    st.text(f"• {f['Archivo']}  →  {f['_motivos']}")
    else:
        st.success(
            f"✅ Los {len(filas_ordenadas)} comprobante(s) se extrajeron sin observaciones."
        )

    df_grilla = pd.DataFrame(
        [
            {
                "Estado": f["Estado"],
                "TIPO": f["TIPO"],
                "CPTE": f["CPTE"],
                "FECHA": f["FECHA"],
                "MES": f["MES"],
                "CANTIDAD KG": f["CANTIDAD KG"],
                "$ BRUTO": f["$ BRUTO"],
                "$ Mercadería": f["$ Mercadería"],
                "$ Comisión": f["$ Comisión"],
                "Filas": f["Filas"],
                "Archivo": f["Archivo"],
            }
            for f in filas_ordenadas
        ]
    )

    st.subheader("Grilla editable")
    st.caption(
        "TIPO, CPTE, FECHA, MES, CANTIDAD KG y $ BRUTO se pueden editar. "
        "Las columnas grises son sólo de auditoría."
    )

    df_editado = st.data_editor(
        df_grilla,
        column_config={
            "Estado": st.column_config.TextColumn("Estado", disabled=True, width="small"),
            "TIPO": st.column_config.SelectboxColumn(
                "TIPO", options=["LCD", "LCDP", "LC"], required=True, width="small"
            ),
            "CPTE": st.column_config.TextColumn("CPTE", required=True, width="medium"),
            "FECHA": st.column_config.DateColumn(
                "FECHA", format="DD/MM/YYYY", width="small"
            ),
            "MES": st.column_config.DateColumn(
                "MES", format="DD/MM/YYYY", width="small"
            ),
            "CANTIDAD KG": st.column_config.NumberColumn(
                "CANTIDAD KG", format="%d", width="small"
            ),
            "$ BRUTO": st.column_config.NumberColumn(
                "$ BRUTO", format="%.2f", width="medium"
            ),
            "$ Mercadería": st.column_config.NumberColumn(
                "$ Mercadería", format="%.2f", disabled=True, width="medium"
            ),
            "$ Comisión": st.column_config.NumberColumn(
                "$ Comisión", format="%.2f", disabled=True, width="small"
            ),
            "Filas": st.column_config.NumberColumn("Filas", disabled=True, width="small"),
            "Archivo": st.column_config.TextColumn("Archivo", disabled=True, width="medium"),
        },
        hide_index=True,
        use_container_width=True,
        key="lc_editor",
        num_rows="fixed",
    )

    # Sincronizar ediciones a session_state
    # Si el usuario edito KG o BRUTO, colapsamos los componentes a un solo valor
    # asi al exportar se escribe el valor editado y no la formula original.
    for i, fila_orig in enumerate(filas_ordenadas):
        idx = next(
            (
                j
                for j, f in enumerate(st.session_state.lc_filas)
                if f["Archivo"] == fila_orig["Archivo"]
            ),
            None,
        )
        if idx is None:
            continue
        f = st.session_state.lc_filas[idx]

        # Campos texto / fecha / select
        f["TIPO"] = df_editado.iloc[i]["TIPO"]
        f["CPTE"] = df_editado.iloc[i]["CPTE"]
        nueva_fecha = df_editado.iloc[i]["FECHA"]
        f["FECHA"] = nueva_fecha if pd.notna(nueva_fecha) else None
        nueva_mes = df_editado.iloc[i]["MES"]
        f["MES"] = nueva_mes if pd.notna(nueva_mes) else None

        # Numericos: si cambian, colapsamos formula
        nuevo_kg = df_editado.iloc[i]["CANTIDAD KG"]
        nuevo_kg_int = int(nuevo_kg) if pd.notna(nuevo_kg) else 0
        if nuevo_kg_int != fila_orig["CANTIDAD KG"]:
            f["_componentes_kg"] = [nuevo_kg_int]
        f["CANTIDAD KG"] = nuevo_kg_int

        nuevo_bruto = df_editado.iloc[i]["$ BRUTO"]
        nuevo_bruto_f = round(float(nuevo_bruto), 2) if pd.notna(nuevo_bruto) else 0.0
        if abs(nuevo_bruto_f - fila_orig["$ BRUTO"]) > 0.01:
            f["_componentes_bruto"] = [nuevo_bruto_f]
        f["$ BRUTO"] = nuevo_bruto_f

    # Export
    st.divider()

    col_a, col_b = st.columns([3, 7])
    with col_a:
        incluir_archivo = st.checkbox(
            "Incluir columna 'Archivo'",
            value=False,
            help="Trazabilidad: agrega el nombre del PDF origen al final.",
        )

    filas_export = sorted(
        st.session_state.lc_filas,
        key=lambda f: (f["FECHA"] or date(1900, 1, 1), f.get("CPTE") or ""),
    )

    excel_bytes = generar_excel(filas_export, incluir_archivo=incluir_archivo)
    nombre_xlsx = (
        f"liquidaciones_carne_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

    st.download_button(
        "📥 Descargar Excel",
        data=excel_bytes,
        file_name=nombre_xlsx,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
else:
    st.info("👆 Subí PDFs y presioná **Procesar PDFs** para empezar.")
